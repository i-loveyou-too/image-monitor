import json
import logging
import os
import shutil
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook

from text_utils import clean_text


INVALID_FS_CHARS = '\\/:*?"<>|'

logger = logging.getLogger(__name__)

# Priority per display field: scraped_* (live DOM extraction) > source_* (Excel import,
# preserved as-is) > pre-existing Seller column > URL-derived last resort.
_SCRAPED_ATTR = {
    "platform": "scraped_platform",
    "category": "scraped_category",
    "seller_name": "scraped_seller_name",
    "product_name": "scraped_product_name",
}
_SOURCE_ATTR = {
    "platform": "source_platform_name",
    "category": "source_category",
    "seller_name": "source_seller_name",
    "product_name": "source_product_name",
}


def build_storage_folder_name(
    *, platform: str | None, category: str | None, seller_name: str | None, product_name: str | None
) -> str:
    parts = [clean_text(platform), clean_text(category), clean_text(seller_name), clean_text(product_name)]
    raw = "_".join(part for part in parts if part) or "unknown"
    return _sanitize_for_windows(raw)


def build_job_folder_name(started_at: datetime) -> str:
    return started_at.strftime("%Y-%m-%d_%H-%M-%S")


def resolve_display(seller, field: str) -> str | None:
    """Priority: scraped_* > source_* > pre-existing Seller field > URL fallback.
    `field` is one of "platform" / "category" / "seller_name" / "product_name"."""
    scraped = getattr(seller, _SCRAPED_ATTR[field], None)
    if scraped:
        return scraped
    source = getattr(seller, _SOURCE_ATTR[field], None)
    if source:
        return source
    if field == "platform" and seller.platform:
        return seller.platform
    if field == "seller_name" and seller.seller_name:
        return seller.seller_name
    if field == "product_name":
        return url_based_fallback_label(seller.product_url)
    return None


def resolve_display_price(seller) -> int | None:
    if seller.scraped_price is not None:
        return seller.scraped_price
    if seller.source_price is not None:
        return seller.source_price
    return None


def url_based_fallback_label(product_url: str | None) -> str | None:
    """Last-resort, URL-derived label used only when no scraped/source/existing
    value is available (e.g. every crawl attempt so far has been blocked)."""
    if not product_url:
        return None
    try:
        parsed = urlparse(product_url)
    except ValueError:
        return None
    query = parse_qs(parsed.query)
    for key in ("prdNo", "productId", "itemId", "item-no"):
        values = query.get(key)
        if values and values[0]:
            return values[0]
    segments = [s for s in parsed.path.split("/") if s]
    if segments:
        return segments[-1]
    return parsed.netloc or None


def resolve_seller_storage_dir(storage_root: str, seller) -> str:
    folder_name = build_storage_folder_name(
        platform=resolve_display(seller, "platform"),
        category=resolve_display(seller, "category"),
        seller_name=resolve_display(seller, "seller_name"),
        product_name=resolve_display(seller, "product_name"),
    )
    base_path = os.path.join(storage_root, folder_name)
    owner_id = _folder_owner_seller_id(base_path)

    if owner_id is None or owner_id == seller.id:
        # Same seller re-crawling (or nobody has claimed this name yet): reuse the folder.
        os.makedirs(base_path, exist_ok=True)
        return base_path

    # A different seller already owns a folder with this exact Platform/Category/
    # SellerName/ProductName combination. Rather than an incrementing _2/_3 suffix
    # (which would fragment THIS seller's history across runs), route this seller to a
    # stable, seller_id-derived folder so every future crawl of this seller lands in the
    # same place.
    conflict_path = f"{base_path}__sid-{seller.id[:8]}"
    conflict_owner_id = _folder_owner_seller_id(conflict_path)
    if conflict_owner_id is not None and conflict_owner_id != seller.id:
        conflict_path = f"{base_path}__sid-{seller.id}"
    logger.warning(
        "storage folder name collision: folder=%r is owned by seller_id=%s; "
        "routing seller_id=%s to %r instead",
        folder_name, owner_id, seller.id, conflict_path,
    )
    os.makedirs(conflict_path, exist_ok=True)
    return conflict_path


def _folder_owner_seller_id(seller_dir: str) -> str | None:
    """Return the seller_id that already owns this folder, based on metadata.json,
    or None if the folder is unclaimed/new. Checks both the new layout (metadata.json
    nested under a per-crawl timestamp subfolder) and the older flat layout
    (metadata.json directly under the seller folder) so pre-existing folders are
    still correctly attributed to their owner."""
    if not os.path.isdir(seller_dir):
        return None
    sid = _read_seller_id_from_metadata(os.path.join(seller_dir, "metadata.json"))
    if sid:
        return sid
    try:
        entries = sorted(os.listdir(seller_dir))
    except OSError:
        return None
    for entry in entries:
        sid = _read_seller_id_from_metadata(os.path.join(seller_dir, entry, "metadata.json"))
        if sid:
            return sid
    return None


def _read_seller_id_from_metadata(meta_path: str) -> str | None:
    if not os.path.isfile(meta_path):
        return None
    try:
        with open(meta_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, ValueError):
        return None
    return data.get("seller_id")


def resolve_job_storage_dir(seller_dir: str, started_at: datetime) -> str:
    return make_unique_dir(os.path.join(seller_dir, build_job_folder_name(started_at)))


def make_unique_dir(base_path: str) -> str:
    if not os.path.exists(base_path):
        _mkdir_or_raise(base_path)
        return base_path
    suffix = 2
    while True:
        candidate = f"{base_path}_{suffix}"
        if not os.path.exists(candidate):
            _mkdir_or_raise(candidate)
            return candidate
        suffix += 1


def write_metadata_json(job_dir: str, *, seller, job) -> str:
    path = os.path.join(job_dir, "metadata.json")
    data = {
        "seller_id": seller.id,
        "crawl_job_id": job.id,
        "platform": seller.platform,
        "product_url": seller.product_url,
        "final_page_url": job.final_page_url,
        "screenshot_path": job.screenshot_path,
        "started_at": job.started_at.isoformat() if job.started_at else None,
        "finished_at": job.finished_at.isoformat() if job.finished_at else None,
        "collected_at": datetime.utcnow().isoformat(),
        "source_platform_name": seller.source_platform_name,
        "source_category": seller.source_category,
        "source_seller_name": seller.source_seller_name,
        "source_product_name": seller.source_product_name,
        "source_price": seller.source_price,
        "scraped_platform": seller.scraped_platform,
        "scraped_category": seller.scraped_category,
        "scraped_seller_name": seller.scraped_seller_name,
        "seller_name_source": seller.seller_name_source,
        "scraped_product_name": seller.scraped_product_name,
        "scraped_price": seller.scraped_price,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


def delete_seller_storage(storage_root: str, seller_id: str):
    legacy_dir = os.path.join(storage_root, seller_id)
    if os.path.isdir(legacy_dir):
        shutil.rmtree(legacy_dir, ignore_errors=True)


def delete_storage_paths(paths: list[str]):
    seller_dirs = set()
    for path in paths:
        if not path:
            continue
        current = Path(path)
        target = current if current.is_dir() else current.parent
        if target.name == "images":
            seller_dirs.add(target.parent.parent)
        elif target.name == "storage":
            continue
        else:
            # page_full.png / metadata.json inside job dir
            seller_dirs.add(target.parent)
    for seller_dir in seller_dirs:
        if seller_dir.exists() and seller_dir.is_dir():
            shutil.rmtree(seller_dir, ignore_errors=True)


def _mkdir_or_raise(path: str):
    try:
        os.makedirs(path, exist_ok=False)
    except OSError as e:
        raise OSError(
            f"Failed to create storage directory '{path}'. "
            "If this is a Windows path-length issue, shorten the storage root path."
        ) from e


def _sanitize_for_windows(value: str) -> str:
    return "".join("_" if ch in INVALID_FS_CHARS else ch for ch in value)


@lru_cache(maxsize=1)
def _excel_rows_by_url() -> dict[str, dict]:
    repo_root = Path(__file__).resolve().parents[1]
    excel_files = list(repo_root.glob("*.xlsx"))
    if not excel_files:
        return {}
    wb = load_workbook(excel_files[0], read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}
    mapping = {}
    for row in rows:
        url = str(row[idx["ProductUrl"]] or "").strip() if "ProductUrl" in idx else ""
        if not url:
            continue
        mapping[url] = {
            "Platform": row[idx["Platform"]] if "Platform" in idx else None,
            "Category": row[idx["Category"]] if "Category" in idx else None,
            "SellerName": row[idx["SellerName"]] if "SellerName" in idx else None,
            "ProductName": row[idx["ProductName"]] if "ProductName" in idx else None,
            "Price": row[idx["Price"]] if "Price" in idx else None,
            # Reference-only: used to visually cross-check scraped images, never
            # treated as a final value.
            "ImageUrl": row[idx["ImageUrl"]] if "ImageUrl" in idx else None,
        }
    return mapping


def lookup_excel_metadata_by_url(product_url: str) -> dict:
    return _excel_rows_by_url().get(product_url, {})
